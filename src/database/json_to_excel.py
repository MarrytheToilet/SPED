#!/usr/bin/env python3
"""
JSON to Excel Exporter
从extracted JSON文件直接导出为Excel多表格式
"""
import json
from pathlib import Path
from typing import Dict, List, Any
from datetime import datetime
from loguru import logger
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter


class JSONToExcelExporter:
    """从JSON直接导出到Excel"""
    
    def __init__(self, 
                 extracted_dir: Path = Path("data/processed/extracted"),
                 schema_file: Path = Path("data_schema/schema.json")):
        """
        初始化导出器
        
        Args:
            extracted_dir: 提取结果JSON文件目录
            schema_file: schema定义文件
        """
        self.extracted_dir = extracted_dir
        self.schema_file = schema_file
        self.schema = self._load_schema()
        
    def _load_schema(self) -> Dict:
        """加载schema定义"""
        try:
            with open(self.schema_file, 'r', encoding='utf-8') as f:
                return json.load(f)
        except Exception as e:
            logger.error(f"加载schema失败: {e}")
            return {"schema_version": "2.0", "tables": []}
    
    def collect_all_records(self) -> List[Dict]:
        """收集所有JSON文件中的记录"""
        all_records = []
        
        if not self.extracted_dir.exists():
            logger.warning(f"提取目录不存在: {self.extracted_dir}")
            return all_records
        
        json_files = list(self.extracted_dir.glob("*.json"))
        logger.info(f"找到 {len(json_files)} 个JSON文件")
        
        for json_file in json_files:
            try:
                with open(json_file, 'r', encoding='utf-8') as f:
                    data = json.load(f)
                
                records = data.get('records', [])
                if records:
                    for record in records:
                        # 提取记录数据
                        record_data = record.get('data', record)
                        if record_data and isinstance(record_data, dict):
                            # 添加元数据
                            record_data['_source_file'] = json_file.name
                            record_data['_paper_id'] = data.get('paper_id', '')
                            all_records.append(record_data)
                            
            except Exception as e:
                logger.error(f"读取文件失败 {json_file.name}: {e}")
                continue
        
        logger.info(f"共收集到 {len(all_records)} 条记录")
        return all_records
    
    def organize_by_tables(self, records: List[Dict]) -> Dict[str, List[Dict]]:
        """按表组织数据"""
        tables_data = {}
        
        # 初始化所有表
        for table_def in self.schema.get('tables', []):
            table_name = table_def.get('table_name')
            if table_name:
                tables_data[table_name] = []
        
        # 填充数据
        for record in records:
            for table_name in tables_data.keys():
                table_data = record.get(table_name)
                if table_data and table_data is not None:
                    # 处理数组格式（如表11可能有多个图片）
                    if isinstance(table_data, list):
                        # 数组格式：每个元素是一条记录
                        for item in table_data:
                            if isinstance(item, dict):
                                row_data = dict(item)
                                row_data['来源文件'] = record.get('_source_file', '')
                                row_data['论文ID'] = record.get('_paper_id', '')
                                tables_data[table_name].append(row_data)
                    elif isinstance(table_data, dict):
                        # 字典格式：单条记录
                        row_data = dict(table_data)
                        row_data['来源文件'] = record.get('_source_file', '')
                        row_data['论文ID'] = record.get('_paper_id', '')
                        tables_data[table_name].append(row_data)
        
        return tables_data
    
    def export_to_excel(self, 
                        output_file: Path,
                        filter_empty: bool = True) -> bool:
        """
        导出为Excel文件
        
        Args:
            output_file: 输出文件路径
            filter_empty: 是否过滤空表
            
        Returns:
            是否成功
        """
        try:
            logger.info("开始收集数据...")
            records = self.collect_all_records()
            
            if not records:
                logger.warning("没有找到任何记录")
                return False
            
            logger.info("按表组织数据...")
            tables_data = self.organize_by_tables(records)
            
            # 创建Excel工作簿
            wb = openpyxl.Workbook()
            wb.remove(wb.active)  # 删除默认sheet
            
            # 按schema顺序创建sheet
            sheet_count = 0
            for table_def in self.schema.get('tables', []):
                table_name = table_def.get('table_name')
                sheet_name = table_def.get('sheet_name', str(sheet_count + 1))
                
                if not table_name:
                    continue
                
                data = tables_data.get(table_name, [])
                
                # 过滤空表
                if filter_empty and not data:
                    logger.debug(f"跳过空表: {table_name}")
                    continue
                
                # 创建sheet
                ws = wb.create_sheet(title=f"Sheet{sheet_name}_{table_name}")
                
                # 写入数据
                self._write_table_to_sheet(ws, table_def, data)
                sheet_count += 1
                
                logger.info(f"✓ {table_name}: {len(data)} 条记录")
            
            if sheet_count == 0:
                logger.warning("没有可导出的数据")
                return False
            
            # 保存文件
            output_file.parent.mkdir(parents=True, exist_ok=True)
            wb.save(output_file)
            
            logger.info(f"导出成功: {output_file}")
            logger.info(f"  - 总表数: {sheet_count}")
            logger.info(f"  - 总记录数: {len(records)}")
            
            return True
            
        except Exception as e:
            logger.error(f"导出失败: {e}", exc_info=True)
            return False
    
    def _write_table_to_sheet(self, 
                               ws, 
                               table_def: Dict, 
                               data: List[Dict]):
        """写入表数据到sheet"""
        # 获取列定义
        columns = table_def.get('columns', [])
        
        # 添加额外列
        extra_columns = [
            {'name': '来源文件', 'description': '数据来源的JSON文件'},
            {'name': '论文ID', 'description': '论文标识'}
        ]
        
        all_columns = columns + extra_columns
        
        # 写表头
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True, size=11)
        header_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        
        for col_idx, col_def in enumerate(all_columns, 1):
            cell = ws.cell(row=1, column=col_idx)
            cell.value = col_def.get('name', '')
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = header_alignment
            
            # 设置列宽
            col_letter = get_column_letter(col_idx)
            ws.column_dimensions[col_letter].width = 20
        
        # 写数据
        for row_idx, row_data in enumerate(data, 2):
            for col_idx, col_def in enumerate(all_columns, 1):
                col_name = col_def.get('name', '')
                value = row_data.get(col_name)
                
                # 处理null值
                if value is None:
                    value = ""
                
                cell = ws.cell(row=row_idx, column=col_idx)
                cell.value = str(value) if value else ""
                cell.alignment = Alignment(vertical='top', wrap_text=True)
        
        # 冻结首行
        ws.freeze_panes = 'A2'


def export_json_to_excel(output_dir: Path = Path("data/exports"),
                         extracted_dir: Path = Path("data/processed/extracted"),
                         schema_file: Path = Path("data_schema/schema.json"),
                         filter_empty: bool = True) -> bool:
    """
    快速导出函数
    
    Args:
        output_dir: 输出目录
        extracted_dir: 提取结果目录
        schema_file: schema文件
        filter_empty: 是否过滤空表
        
    Returns:
        是否成功
    """
    logger.info("="*80)
    logger.info("从JSON直接导出Excel")
    logger.info("="*80)
    
    exporter = JSONToExcelExporter(extracted_dir, schema_file)
    
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    output_file = output_dir / f"人工关节数据_{timestamp}.xlsx"
    
    success = exporter.export_to_excel(output_file, filter_empty)
    
    if success:
        logger.info("="*80)
        logger.info(f"导出完成: {output_file}")
        logger.info("="*80)
    
    return success


if __name__ == "__main__":
    # 测试
    import sys
    from loguru import logger
    
    logger.remove()
    logger.add(sys.stderr, level="INFO")
    
    export_json_to_excel()
