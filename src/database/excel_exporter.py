#!/usr/bin/env python3
"""
Excel多表导出器 - 按照新schema(v3.0)导出为多sheet的Excel文件
"""
import json
import sqlite3
from pathlib import Path
from typing import Dict, List, Any, Optional
from datetime import datetime
from loguru import logger
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

from .db_manager import DatabaseManager


class ExcelExporter:
    """Excel多表导出类"""
    
    # 中文表名到数据库表名的映射
    TABLE_NAME_MAPPING = {
        'Sheet1_基本信息表': 'basic_info',
        'Sheet2_内衬基本信息表': 'liner_basic',
        'Sheet3_球头基本信息表': 'head_basic',
        'Sheet4_配合信息表': 'fitting_info',
        'Sheet5_股骨柄基本信息表': 'stem_basic',
        'Sheet6_内衬物理性能表': 'liner_properties',
        'Sheet7_球头物理性能表': 'head_properties',
        'Sheet8_股骨柄物理性能表': 'stem_properties',
        'Sheet9_实验参数': 'experiment_params',
        'Sheet10_性能测试结果表': 'test_results',
        'Sheet11_计算模拟参数表': 'simulation_params',
        'Sheet12_计算模拟图像表': 'simulation_images'
    }
    
    def __init__(self, 
                 db_manager: Optional[DatabaseManager] = None,
                 schema_file: Path = Path("data_schema/schema.json")):
        """
        初始化Excel导出器
        
        Args:
            db_manager: 数据库管理器实例
            schema_file: schema文件路径
        """
        self.db = db_manager or DatabaseManager()
        self.schema_file = schema_file
        self.schema = self._load_schema()
    
    def _load_schema(self) -> Dict:
        """加载schema定义"""
        try:
            with open(self.schema_file, 'r', encoding='utf-8') as f:
                return json.load(f)
        except Exception as e:
            logger.error(f"加载schema失败: {e}")
            return {"schema_version": "3.0", "tables": []}
    
    def export_by_schema(self, output_file: Path, filter_empty_sheets: bool = True) -> bool:
        """
        按照schema定义导出多表Excel
        
        Args:
            output_file: 输出文件路径
            filter_empty_sheets: 是否过滤空sheet
            
        Returns:
            是否成功
        """
        try:
            logger.info(f"开始导出Excel: {output_file}")
            
            # 创建工作簿
            wb = openpyxl.Workbook()
            wb.remove(wb.active)  # 删除默认sheet
            
            # 遍历schema中定义的每个表
            tables = self.schema.get('tables', [])
            if not tables:
                logger.error("Schema中没有定义表")
                return False
            
            sheet_count = 0
            for table_def in tables:
                sheet_name = table_def.get('table_name', '')
                db_table_name = table_def.get('db_table_name', self.TABLE_NAME_MAPPING.get(sheet_name, ''))
                
                # 从数据库获取数据
                records = self.db.get_all_records(db_table_name)
                
                # 如果过滤空sheet，检查是否所有记录都是空的（只有数据ID和时间戳）
                if filter_empty_sheets:
                    # 检查是否有实际数据
                    has_data = False
                    for record in records:
                        # 检查除了数据ID、created_at、updated_at之外是否有数据
                        for key, value in record.items():
                            if key not in ['数据ID', 'created_at', 'updated_at'] and value and value != '':
                                has_data = True
                                break
                        if has_data:
                            break
                    
                    if not has_data:
                        logger.debug(f"跳过空表: {sheet_name}（无实际数据）")
                        continue
                
                # 创建sheet
                ws = wb.create_sheet(title=sheet_name)
                
                # 获取列定义
                columns = table_def.get('columns', [])
                if not columns:
                    logger.warning(f"表 {sheet_name} 没有列定义")
                    continue
                
                # 写入表头
                header_row = [col['name'] for col in columns]
                ws.append(header_row)
                
                # 美化表头
                for col_idx, _ in enumerate(header_row, 1):
                    cell = ws.cell(row=1, column=col_idx)
                    cell.font = Font(bold=True, color='FFFFFF')
                    cell.fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
                    cell.alignment = Alignment(horizontal='center', vertical='center')
                
                # 写入数据
                for record in records:
                    row_data = []
                    for col in columns:
                        col_name = col['name']
                        value = record.get(col_name, '')
                        
                        # 处理None值
                        if value is None or value == 'null':
                            row_data.append('')
                        else:
                            row_data.append(str(value))
                    
                    ws.append(row_data)
                
                # 自动调整列宽
                for col_idx in range(1, len(header_row) + 1):
                    col_letter = get_column_letter(col_idx)
                    max_length = 0
                    
                    for row in ws[col_letter]:
                        try:
                            if len(str(row.value)) > max_length:
                                max_length = len(str(row.value))
                        except:
                            pass
                    
                    adjusted_width = min(max_length + 2, 50)
                    ws.column_dimensions[col_letter].width = adjusted_width
                
                sheet_count += 1
                logger.debug(f"✓ 导出 {sheet_name}: {len(records)} 条记录")
            
            # 保存文件
            if sheet_count == 0:
                logger.warning("没有数据可导出")
                return False
            
            wb.save(output_file)
            logger.info(f"✅ Excel导出成功: {output_file} ({sheet_count} 个sheets)")
            return True
            
        except Exception as e:
            logger.error(f"导出Excel失败: {e}")
            import traceback
            logger.error(traceback.format_exc())
            return False
    
    def export_simple(self, output_file: Path) -> bool:
        """
        简单导出（不依赖schema，直接从数据库导出）
        
        Args:
            output_file: 输出文件路径
            
        Returns:
            是否成功
        """
        try:
            wb = openpyxl.Workbook()
            wb.remove(wb.active)
            
            for sheet_name, table_name in self.TABLE_NAME_MAPPING.items():
                records = self.db.get_all_records(table_name)
                
                if not records:
                    continue
                
                ws = wb.create_sheet(title=sheet_name)
                
                # 使用第一条记录的键作为表头
                headers = list(records[0].keys())
                ws.append(headers)
                
                # 美化表头
                for col_idx in range(1, len(headers) + 1):
                    cell = ws.cell(row=1, column=col_idx)
                    cell.font = Font(bold=True, color='FFFFFF')
                    cell.fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
                    cell.alignment = Alignment(horizontal='center', vertical='center')
                
                # 写入数据
                for record in records:
                    row_data = [record.get(h, '') for h in headers]
                    ws.append(row_data)
                
                # 自动调整列宽
                for col_idx in range(1, len(headers) + 1):
                    ws.column_dimensions[get_column_letter(col_idx)].width = 20
            
            wb.save(output_file)
            logger.info(f"✅ Excel导出成功: {output_file}")
            return True
            
        except Exception as e:
            logger.error(f"导出Excel失败: {e}")
            return False


def export_to_excel(output_dir: Path = Path("data/exports"), 
                    filter_empty: bool = True) -> Optional[Path]:
    """
    导出所有表到Excel（按schema组织）
    
    Args:
        output_dir: 输出目录
        filter_empty: 是否过滤空表
        
    Returns:
        输出文件路径，失败返回None
    """
    try:
        output_dir.mkdir(parents=True, exist_ok=True)
        
        # 生成文件名（带时间戳）
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        output_file = output_dir / f"artificial_joint_data_{timestamp}.xlsx"
        
        logger.info("开始导出Excel多表文件（按schema组织）")
        
        exporter = ExcelExporter()
        success = exporter.export_by_schema(output_file, filter_empty_sheets=filter_empty)
        
        if success:
            logger.info(f"✅ 导出完成: {output_file}")
            return output_file
        else:
            logger.error("导出失败")
            return None
            
    except Exception as e:
        logger.error(f"导出过程出错: {e}")
        return None


if __name__ == "__main__":
    # 测试导出
    result = export_to_excel()
    if result:
        print(f"✅ 导出成功: {result}")
    else:
        print("❌ 导出失败")
